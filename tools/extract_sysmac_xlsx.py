from pathlib import Path
from openpyxl import load_workbook
import json
import re


BASE_DIR = Path(__file__).resolve().parent.parent
RAW_DIR = BASE_DIR / "company_template" / "raw_xlsx"
ST_DIR = BASE_DIR / "company_template" / "st"
META_DIR = BASE_DIR / "company_template" / "meta"

ST_DIR.mkdir(parents=True, exist_ok=True)
META_DIR.mkdir(parents=True, exist_ok=True)


def safe_name(name: str) -> str:
    name = name.replace("(1)", "")
    name = name.replace("_程序", "")
    name = name.replace("程序", "")
    name = name.replace(" ", "_")
    name = name.replace("（", "_").replace("）", "")
    name = name.replace("(", "_").replace(")", "")
    name = re.sub(r"_+", "_", name)
    return name.strip("_")


def clean_cell(value) -> str:
    if value is None:
        return ""

    text = str(value)

    # 去掉 Excel 单元格里的多余空白
    text = text.replace("\r\n", "\n")
    text = text.replace("\r", "\n")

    return text.strip()


def extract_sheet_text(ws) -> str:
    """
    把一个 sheet 中的所有非空单元格按行提取出来。
    Sysmac 导出的 Excel 有时候 ST 不一定固定在 ST sheet，
    所以这里做成通用提取。
    """
    lines = []

    for row in ws.iter_rows(values_only=True):
        row_texts = []

        for cell in row:
            text = clean_cell(cell)
            if text:
                row_texts.append(text)

        if row_texts:
            # 如果一行有多个单元格，用 tab 隔开
            lines.append("\t".join(row_texts))

    return "\n".join(lines).strip()


def looks_like_st(text: str) -> bool:
    """
    判断文本是否像 ST 程序。
    """
    keywords = [
        "PROGRAM",
        "END_PROGRAM",
        "FUNCTION_BLOCK",
        "END_FUNCTION_BLOCK",
        "VAR",
        "END_VAR",
        "IF",
        "END_IF",
        "CASE",
        "END_CASE",
        ":=",
    ]

    score = 0
    upper_text = text.upper()

    for kw in keywords:
        if kw in upper_text:
            score += 1

    return score >= 2


def extract_best_st(wb) -> tuple[str, dict]:
    """
    尽量找到最像 ST 的 sheet。
    同时导出每个 sheet 的全文到 meta 方便排查。
    """
    sheet_dumps = {}
    candidates = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        text = extract_sheet_text(ws)
        sheet_dumps[sheet_name] = text

        if not text:
            continue

        # 优先 sheet 名包含这些关键词
        name_score = 0
        if "ST" in sheet_name.upper():
            name_score += 5
        if "程序" in sheet_name:
            name_score += 3
        if "代码" in sheet_name:
            name_score += 3

        st_score = 0
        upper_text = text.upper()

        for kw in [
            "PROGRAM",
            "END_PROGRAM",
            "FUNCTION_BLOCK",
            "END_FUNCTION_BLOCK",
            "VAR",
            "END_VAR",
            "IF",
            "END_IF",
            "CASE",
            "END_CASE",
            ":=",
        ]:
            if kw in upper_text:
                st_score += 1

        total_score = name_score + st_score

        candidates.append({
            "sheet_name": sheet_name,
            "score": total_score,
            "text": text,
        })

    if not candidates:
        return "", sheet_dumps

    candidates.sort(key=lambda x: x["score"], reverse=True)
    best = candidates[0]

    # 如果最高分太低，也仍然导出，方便人工检查
    return best["text"], sheet_dumps


def main():
    files = list(RAW_DIR.glob("*.xlsx"))

    if not files:
        print(f"没有找到 xlsx 文件：{RAW_DIR}")
        return

    for file in files:
        print(f"处理：{file.name}")

        try:
            wb = load_workbook(file, data_only=False)
        except Exception as e:
            print(f"  打开失败：{e}")
            continue

        base_name = safe_name(file.stem)

        st_text, sheet_dumps = extract_best_st(wb)

        # 保存最像 ST 的内容
        if st_text:
            st_path = ST_DIR / f"{base_name}.st"
            st_path.write_text(st_text, encoding="utf-8")
            print(f"  ST 已导出：{st_path}")
        else:
            print("  未提取到 ST 内容")

        # 保存每个 sheet 的原始文本，方便后面排查
        meta_path = META_DIR / f"{base_name}_sheets.json"
        meta_path.write_text(
            json.dumps(sheet_dumps, ensure_ascii=False, indent=4),
            encoding="utf-8"
        )
        print(f"  Sheet文本已导出：{meta_path}")

    print("")
    print("模板提取完成。")


if __name__ == "__main__":
    main()